#!/usr/bin/env python3
"""
Build an HTML + Excel report from perf stat outputs.

By default: uses the LATEST timestamp under each results/<variant>/*/perf/
Optionally: pick a SPECIFIC timestamp for all variants with --timestamp YYYYMMDD_HHMMSS

Outputs:
  reports/<run_ts>/perf_report.html
  reports/<run_ts>/perf_report.xlsx
"""

import argparse
import math
import re
from collections import defaultdict, OrderedDict
from datetime import datetime
from pathlib import Path

import pandas as pd
import plotly.graph_objects as go

ROOT_DEFAULT = Path("results")
REPORT_ROOT_DEFAULT = Path("reports")

COUNTER_KEYS = [
    "time",                 # seconds
    "instructions",
    "cycles",
    "IPC",

    "context-switches",
    "page-faults",

    "branches",
    "branch-misses",

    "L1-dcache-loads",
    "L1-dcache-load-misses",

    "LLC-loads",
    "LLC-load-misses",


    "dTLB-loads",
    "dTLB-load-misses",

    "iTLB-loads",
    "iTLB-load-misses",

    "Speedup",
]

LABEL_NORMALIZE = {
    "cycles": "cycles",
    "instructions": "instructions",

    "context-switches": "context-switches",
    "page-faults": "page-faults",

    "branches": "branches",
    "branch-misses": "branch-misses",

    "l1-dcache-loads": "L1-dcache-loads",
    "l1-dcache-load-misses": "L1-dcache-load-misses",

    "llc-loads": "LLC-loads",
    "llc-load-misses": "LLC-load-misses",


    "dtlb-loads": "dTLB-loads",
    "dtlb-load-misses": "dTLB-load-misses",

    "itlb-loads": "iTLB-loads",
    "itlb-load-misses": "iTLB-load-misses",

    # time sources
    "time elapsed": "time",         # seconds time elapsed (if present)
}

NUM_RE = re.compile(r"""
    (?P<num>
        (?:\d{1,3}(?:[,\s]\d{3})+|\d+)
        (?:\.\d+)? | \d+\.\d+
    )
""", re.VERBOSE)

# ---------------------- parsing helpers ----------------------

def _to_number(tok: str):
    tok = tok.replace(" ", "").replace(",", "")
    try:
        return float(tok)
    except Exception:
        return None

def parse_perf_file(path: Path) -> dict:
    results = {}
    text = path.read_text(errors="ignore")
    for line in text.splitlines():
        s = line.strip()
        if not s or "<not supported>" in s.lower():
            continue
        m = NUM_RE.search(s)
        if not m:
            continue
        val = _to_number(m.group("num"))
        if val is None:
            continue

        tail = s[m.end():].strip()
        tail = tail.split("#", 1)[0].strip()
        lbl_low = tail.lower()

        norm = None
        for k, v in LABEL_NORMALIZE.items():
            if lbl_low.endswith(k):
                norm = v
                break
        if norm is None:
            parts = tail.split()
            if parts:
                norm = LABEL_NORMALIZE.get(parts[-1].lower(), LABEL_NORMALIZE.get(tail.lower()))
        if not norm:
            continue

        if norm == "time":
            if "seconds time elapsed" in lbl_low or "time elapsed" in lbl_low:
                results["time"] = val
            else:
                results.setdefault("time", val)
        else:
            results[norm] = val
    return results

def aggregate_variant(perf_dir: Path, geo_mean: bool) -> dict:
    files = sorted(perf_dir.glob("perf_run_*.txt"))
    if not files:
        return {}
    sums = defaultdict(float)
    counts = defaultdict(int)
    for f in files:
        vals = parse_perf_file(f)
        for k, v in vals.items():
            if v is None or (isinstance(v, float) and (math.isnan(v) or math.isinf(v))):
                continue
            sums[k] += v
            counts[k] += 1
    if not geo_mean:
        avg = {k: (sums[k] / counts[k]) for k in sums if counts[k]}
        if avg.get("cycles", 0) > 0 and "instructions" in avg:
            avg["IPC"] = avg["instructions"] / avg["cycles"]
        return avg
    else:
        import numpy as np

        # Collect values per key
        all_vals = defaultdict(list)
        for f in files:
            vals = parse_perf_file(f)
            for k, v in vals.items():
                if v is None or (isinstance(v, float) and (math.isnan(v) or math.isinf(v))):
                    continue
                all_vals[k].append(v)

        # Compute geometric mean for each metric
        geo_avg = {}
        for k, arr in all_vals.items():
            vals = [v for v in arr if v > 0]  # geometric mean only defined for positive values
            if not vals:
                continue
            geo_avg[k] = float(np.exp(np.mean(np.log(vals))))

        # Derived metric: IPC
        if geo_avg.get("cycles", 0) > 0 and "instructions" in geo_avg:
            geo_avg["IPC"] = geo_avg["instructions"] / geo_avg["cycles"]

        return geo_avg


# ---------------------- discovery ----------------------

def latest_timestamp_dir(variant_dir: Path) -> Path | None:
    stamps = sorted([p for p in variant_dir.iterdir() if p.is_dir()])
    return stamps[-1] if stamps else None

def perf_dir_for_variant(variant_dir: Path, forced_timestamp: str | None) -> Path | None:
    """
    If forced_timestamp is given, use results/<variant>/<forced_timestamp>/perf
    Else, use latest timestamp under results/<variant>/*/perf
    """
    if forced_timestamp:
        p = variant_dir / forced_timestamp / "perf"
        return p if p.exists() else None
    stamp_dir = latest_timestamp_dir(variant_dir)
    if not stamp_dir:
        return None
    p = stamp_dir / "perf"
    return p if p.exists() else None

def find_variants_perf_dirs(root: Path, forced_timestamp: str | None) -> dict:
    """
    Locate perf directories for all variants.
    Supports structures like:
      <root>/<variant>/<timestamp>/perf/
    """
    out = {}
    if not root.exists():
        return out

    for variant_dir in sorted([p for p in root.iterdir() if p.is_dir()]):
        # Check all timestamped dirs inside variant_dir
        candidates = []
        for ts_dir in sorted(variant_dir.iterdir()):
            if not ts_dir.is_dir():
                continue
            perf_dir = ts_dir / "perf"
            if perf_dir.exists():
                candidates.append((ts_dir.name, perf_dir))

        # Pick forced timestamp or latest one
        chosen = None
        if forced_timestamp:
            for ts, p in candidates:
                if ts == forced_timestamp:
                    chosen = p
                    break
        elif candidates:
            chosen = candidates[-1][1]  # latest timestamp

        if chosen:
            out[variant_dir.name] = chosen

    return out

# ---------------------- reporting ----------------------

def make_dataframe(variant_to_avgs: dict, baseline_variant: str | None) -> pd.DataFrame:
    rows = []
    for variant, data in variant_to_avgs.items():
        row = {"variant": variant}
        for k in COUNTER_KEYS:
            if k in ("IPC", "Speedup"):
                continue
            row[k] = data.get(k, float("nan"))
        if "IPC" in data:
            row["IPC"] = data["IPC"]
        rows.append(row)

    df = pd.DataFrame(rows).set_index("variant")
    baseline = None
    if baseline_variant and baseline_variant in df.index:
        baseline = baseline_variant
    else:
        # try to find one ending with '_clean'
        clean_variants = [v for v in df.index if v.lower().endswith("_clean")]
        if clean_variants:
            baseline = clean_variants[0]
        elif len(df.index):
            baseline = df.index[0]

    if baseline and pd.notna(df.loc[baseline].get("time", float("nan"))) and df.loc[baseline, "time"] > 0:
        base_time = df.loc[baseline, "time"]
        df["Speedup"] = base_time / df["time"]
    else:
        df["Speedup"] = float("nan")
    # --- Round selected integer metrics ---
    for col in df.columns:
        if col not in ("time", "IPC", "Speedup"):
            df[col] = df[col].apply(lambda x: int(round(x)) if pd.notna(x) else x)

    # --- Round selected float metrics ---
    for col in ("time", "IPC", "Speedup"):
        if col in df.columns:
            df[col] = df[col].round(2)

    # --- Sort columns (i.e., variants after transpose) by Speedup ---
    if "Speedup" in df.columns:
        # Get speedup values per variant
        speedups = df["Speedup"].sort_values(ascending=True)
        # Reorder DataFrame rows (which will become columns after transpose)
        df = df.loc[speedups.index]

    # --- Column order consistency ---
    ordered = [c for c in COUNTER_KEYS if c in df.columns]
    return df[ordered]

def fmt_num(x):
    if pd.isna(x):
        return ""
    if float(x).is_integer():
        return f"{int(x):,}"
    return f"{x:,.3f}"

from openpyxl.styles import numbers

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def write_excel(df_table: pd.DataFrame, out_xlsx: Path):
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as w:
        df_table.to_excel(w, sheet_name="perf_avg", index=True)
        ws = w.book["perf_avg"]
        ws.freeze_panes = "B2"

        # --- Styles ---
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        alt_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        border_style = Side(style="thin", color="CCCCCC")

        # --- Header styling ---
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(top=border_style, bottom=border_style, left=border_style, right=border_style)

        # --- Body styling ---
        for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                val = cell.value
                # Alternate row background
                if r_idx % 2 == 0:
                    cell.fill = alt_fill

                # Border
                cell.border = Border(top=border_style, bottom=border_style, left=border_style, right=border_style)

                # Alignment
                cell.alignment = Alignment(horizontal="right", vertical="center")

                # Number formatting
                if isinstance(val, (int, float)):
                    if abs(val) >= 1e8:
                        cell.number_format = '#,##0'
                    else:
                        cell.number_format = '0.00'

        # --- Auto column widths ---
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = 0
            for c in col_cells:
                v = str(c.value) if c.value is not None else ""
                max_len = max(max_len, len(v))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len + 2), 40)



def bar(fig, title, x, y):
    fig.add_trace(go.Bar(x=x, y=y, text=[fmt_num(v) for v in y], textposition="auto"))
    fig.update_layout(title=title, bargap=0.25)

def write_html(df_table: pd.DataFrame, df_charts: pd.DataFrame, out_html: Path, header_note: str):
    out_html.parent.mkdir(parents=True, exist_ok=True)
    parts = []
    parts.append(f"<h1>perf report</h1>")
    parts.append(f"<p>Generated: {datetime.now():%Y-%m-%d %H:%M:%S} &nbsp;&nbsp; {header_note}</p>")
    parts.append("<h2>Averages</h2>")
    parts.append(df_table.to_html(classes='table', escape=False, justify='center',
                                  formatters={c: fmt_num for c in df_table.columns}))
    parts.append("<hr/><h2>Counters</h2>")

    # charts use the original (non-transposed) orientation: variants on X
    variants = df_charts.index.tolist()
    for col in df_charts.columns:
        s = df_charts[col]
        if s.isna().all():
            continue
        fig = go.Figure()
        fig.add_trace(go.Bar(x=variants, y=s.fillna(0.0).tolist(),
                             text=[fmt_num(v) for v in s.fillna(0.0).tolist()],
                             textposition="auto"))
        fig.update_layout(title=col, bargap=0.25)
        parts.append(fig.to_html(full_html=False, include_plotlyjs="cdn"))

    html = f"""<!doctype html>
<html><head><meta charset="utf-8"><title>perf report</title>
<style>
body {{ font-family: system-ui, sans-serif; margin: 24px; }}
.table {{ border-collapse: collapse; margin-bottom: 24px; }}
.table th, .table td {{ border: 1px solid #ddd; padding: 6px 10px; text-align: right; }}
.table th:first-child, .table td:first-child {{ text-align: left; }}
h1, h2 {{ margin: 8px 0 12px; }}
</style></head><body>
{''.join(parts)}
</body></html>"""
    out_html.write_text(html, encoding="utf-8")

# ---------------------- CLI ----------------------

def parse_args():
    p = argparse.ArgumentParser(description="Build perf HTML + Excel report from perf stat outputs.")
    p.add_argument("--results-dir", default=str(ROOT_DEFAULT), help="Path to results/ (default: results)")
    p.add_argument("--report-dir", default=str(REPORT_ROOT_DEFAULT), help="Where to write reports/ (default: reports)")
    p.add_argument("--timestamp", help="Use this timestamp under each variant (YYYYMMDD_HHMMSS). If absent, use latest per-variant.")
    p.add_argument("--baseline", help="Variant name to use as speedup baseline.")
    p.add_argument("--transpose", action="store_true",
               help="Transpose the data table in Excel and HTML (charts remain the same).")
    p.add_argument("--geomean", action="store_true",
                help="Append geometric mean of Speedup across variants to the report.")

    return p.parse_args()

def main():
    args = parse_args()
    root = Path(args.results_dir)
    report_root = Path(args.report_dir)
    forced_ts = args.timestamp

    found = find_variants_perf_dirs(root, forced_ts)
    if not found:
        print(f"❌ No perf dirs found under {root}/<variant>/{forced_ts or '*'} /perf")
        return

    # Aggregate
    variant_to_avgs = OrderedDict()
    for variant, perf_dir in sorted(found.items()):
        avg = aggregate_variant(perf_dir, args.geomean)
        if not avg:
            print(f"⚠️  No counters parsed in {perf_dir}")
        variant_to_avgs[variant] = avg

    df = make_dataframe(variant_to_avgs, args.baseline)

    df_table = df.T if args.transpose else df
    df_charts = df  # charts keep variants on X

    run_stamp = forced_ts
    benchmark_prefix = Path(args.results_dir).name.split("_")[0]
    result_agg = "geomean" if args.geomean else "arithmetic"
    out_dir = report_root / f"{benchmark_prefix}_results_{result_agg}_{run_stamp}"
    out_html = out_dir / "perf_report.html"
    out_xlsx = out_dir / "perf_report.xlsx"

    write_excel(df_table, out_xlsx)
    header_note = (f"source timestamp: {forced_ts}" if forced_ts
                else "source timestamp: latest per variant")
    header_note += f" · aggregation: {'geometric mean' if args.geomean else 'arithmetic mean'}"
    write_html(df_table, df_charts, out_html, header_note)

    print("\n=== REPORT BUILT ===")
    print(f"HTML : {out_html.resolve()}")
    print(f"Excel: {out_xlsx.resolve()}\n")
    with pd.option_context("display.max_columns", None, "display.width", 160):
        print(df.applymap(fmt_num))

if __name__ == "__main__":
    main()
